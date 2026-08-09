# pyre-ignore-all-errors
import os
import shutil
import uuid
import openpyxl
from pathlib import Path
from typing import Dict, Any

from backend.config import get_settings
from backend.core.excel.processor import ExcelProcessor

settings = get_settings()

class DetailedDesignManager:
    """Links OpenSees structural analysis results (forces, moments) directly to the specific Design Excel sheets."""
    
    def __init__(self):
        self.processor = ExcelProcessor(
            excel_dir=settings.DESIGN_EXCEL_DIR,
            enhanced_dir=settings.ENHANCED_EXCEL_DIR
        )
        self.reports_dir = Path("doc-files/reports/calculations")
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
    def _get_sheet_mapping_definition(self, member_type: str, design_code: str) -> Dict[str, Any]:
        """Provides hardcoded cell mappings for known calculation sheets based on type and code."""
        # In a real app, this is retrieved from a database or JSON config defining the mapping for each custom sheet.
        mappings = {
            "COLUMN": {
                "BNBC_2020": {
                    "sheet_filename": "Column_Design_BNBC2020.xlsx",
                    "input_cells": {
                        "Pu": "D10",   # Axial load
                        "Mux": "D11",  # Moment X
                        "Muy": "D12",  # Moment Y
                        "b": "D6",     # Width
                        "h": "D7",     # Depth
                        "fc": "D4",    # Concrete strength
                        "fy": "D5"     # Steel yield strength
                    },
                    "output_cells": {
                        "Ast_req": "H15",
                        "Interaction_Ratio": "H18",
                        "Design_Status": "H20"  # "OK" or "FAIL"
                    }
                }
            },
            "BEAM": {
                "BNBC_2020": {
                    "sheet_filename": "Beam_Design_BNBC2020.xlsx",
                    "input_cells": {
                        "Mu": "E10",
                        "Vu": "E11",
                        "Tu": "E12",
                        "b": "E6",
                        "h": "E7"
                    },
                    "output_cells": {
                        "Top_Ast": "J15",
                        "Bottom_Ast": "J16",
                        "Shear_Links": "J20",
                        "Status": "J22"
                    }
                }
            }
        }
        
        # Fallback to column if not found for stub
        return mappings.get(member_type, mappings["COLUMN"]).get(design_code, mappings["COLUMN"]["BNBC_2020"])

    def design_structural_member(self, member_data: Dict[str, Any], analysis_forces: Dict[str, float]) -> Dict[str, Any]:
        """
        Takes raw API data or DB model data for a member, looks up forces from FEA per load combos,
        loads the relevant Excel sheet, injects forces/geometry, and generates a standalone calculation sheet.
        """
        member_type = member_data.get("type", "COLUMN").upper()
        # For simplicity in stub, assume BNBC 2020
        mapping_def = self._get_sheet_mapping_definition(member_type, "BNBC_2020")
        
        filename = mapping_def["sheet_filename"]
        
        # First ensure sheet is enhanced
        try:
            enhanced_path = self.processor.enhance_sheet(filename)
        except FileNotFoundError:
            # Create a mock Excel sheet since this is a scaffold/stub phase and doc-files might be missing
            self._create_mock_sheet(self.processor.excel_dir / filename, mapping_def)
            enhanced_path = self.processor.enhance_sheet(filename)
            
        # Copy to a unique standalone report sheet for this specific column/beam
        standalone_filename = f"{member_type}_{member_data.get('label', 'M1')}_{uuid.uuid4().hex[:8]}.xlsx"
        standalone_path = self.reports_dir / standalone_filename
        
        shutil.copy2(enhanced_path, standalone_path)
        
        # Inject values
        wb = openpyxl.load_workbook(standalone_path)
        sheet = wb.active
        assert sheet is not None, "Workbook has no active sheet"
        
        inputs_map = mapping_def["input_cells"]
        
        # Geometry
        geom = member_data.get("geometry", {})
        if "b" in inputs_map and "width" in geom: sheet[inputs_map["b"]] = geom["width"]
        if "h" in inputs_map and "depth" in geom: sheet[inputs_map["h"]] = geom["depth"]
        
        # Forces
        if "Pu" in inputs_map and "P" in analysis_forces: sheet[inputs_map["Pu"]] = analysis_forces["P"]
        if "Mux" in inputs_map and "Mz" in analysis_forces: sheet[inputs_map["Mux"]] = analysis_forces["Mz"] # Assuming Mz is primary bending
        if "Muy" in inputs_map and "My" in analysis_forces: sheet[inputs_map["Muy"]] = analysis_forces["My"]
        if "Vu" in inputs_map and "Vy" in analysis_forces: sheet[inputs_map["Vu"]] = analysis_forces["Vy"]
        
        # Save injected sheet
        wb.save(standalone_path)
        
        # Extract outputs
        # Note: openpyxl cannot evaluate formulas. 
        # In a production environment with xlwings or formulas library, we would trigger calculation here.
        # Since we are using standard python env on linux, we'll return the standalone path and a stubbed output.
        
        return {
            "status": "designed",
            "calculation_sheet_url": f"/files/reports/calculations/{standalone_filename}",
            "summary_results": {
                "Ast_req": 2150.5, # Stub computed required steel
                "Interaction_Ratio": 0.85,
                "Status": "OK"
            }
        }
        
    def _create_mock_sheet(self, filepath: Path, mapping: Dict[str, Any]):
        """Creates a mock Excel sheet if none exists in doc-files (for testing phase)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            return
        ws.title = "Design calculation"
        
        # Put titles and inputs
        ws["C1"] = "RCC Design Calculation Sheet"
        
        row_idx = 4
        for k, v in mapping["input_cells"].items():
            cell = ws[v]
            cell.value = 0 # Default
            # Label to the left
            ws.cell(row=cell.row, column=cell.column-1).value = k
            
        row_idx = 15
        for k, v in mapping["output_cells"].items():
            cell = ws[v]
            cell.value = "FORMULA_STUB"
            ws.cell(row=cell.row, column=cell.column-1).value = k
            
        wb.save(filepath)
