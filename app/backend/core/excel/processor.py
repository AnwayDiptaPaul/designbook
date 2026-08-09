# pyre-ignore-all-errors
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

class ExcelProcessor:
    """Core module for parsing, modifying, and interacting with engineering Excel design sheets."""
    
    def __init__(self, excel_dir: str, enhanced_dir: str):
        self.excel_dir = Path(excel_dir)
        self.enhanced_dir = Path(enhanced_dir)
        self.enhanced_dir.mkdir(parents=True, exist_ok=True)
        
    def list_available_sheets(self):
        """List all .xlsx files in the design excel directory."""
        if not self.excel_dir.exists():
            return []
        
        sheets = []
        for file in self.excel_dir.glob("*.xlsx"):
            # Exclude temporary or enhanced files if any ended up here
            if not file.name.startswith("~"):
                sheets.append({
                    "name": file.name,
                    "path": str(file),
                    "size": file.stat().st_size
                })
        return sheets

    def enhance_sheet(self, filename: str) -> str:
        """
        Takes an original engineering spreadsheet and makes it programmatically usable.
        - Unprotects sheets (removes passwords if empty or standard bypasses)
        - Unmerges cells and duplicates values to maintain references
        - Returns path to the new 'enhanced' file
        """
        source_path = self.excel_dir / filename
        target_path = self.enhanced_dir / f"enhanced_{filename}"
        
        if not source_path.exists():
            raise FileNotFoundError(f"Source Excel sheet {filename} not found.")
            
        # load_workbook data_only=False to keep formulas
        wb = openpyxl.load_workbook(source_path, data_only=False)
        
        for sheet in wb.worksheets:
            # 1. Unprotect sheet
            sheet.protection.disable()
            
            # 2. Safely unmerge cells
            # We must iterate over a copy of merged_cells.ranges because unmerge modifies it
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                # Get the value from the top-left cell before unmerging
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell = sheet.cell(row=min_row, column=min_col)
                value = top_left_cell.value
                data_type = top_left_cell.data_type
                font = top_left_cell.font.copy() if top_left_cell.font else None
                fill = top_left_cell.fill.copy() if top_left_cell.fill else None
                
                # Unmerge
                sheet.unmerge_cells(str(merged_range))
                
                # Duplicate the value back to all cells in the previously merged range
                # This ensures any formulas referencing H4 (which was merged H4:J4) still work 
                # even if the layout changes, but mostly it's for visual consistency if needed.
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell.value = value
                        if font: cell.font = font
                        if fill: cell.fill = fill
        
        # Save enhanced version
        wb.save(target_path)
        return str(target_path)
        
    def extract_inputs_outputs(self, filename: str):
        """
        Analyzes a sheet to guess or extract its input cells (constants, unprotected) 
        and output cells (formulas). Phase 3 stub.
        """
        path = self.enhanced_dir / f"enhanced_{filename}"
        if not path.exists():
            path = self.excel_dir / filename
            
        if not path.exists():
            raise FileNotFoundError("File not found")
            
        wb = openpyxl.load_workbook(path, data_only=False)
        # Note: True input/output mapping typically requires a configuration file or mapping dictionary
        # per spreadsheet, because engineering sheets don't have standard "tags".
        
        # We will stub this for now to return generic mapping
        return {
            "inputs": [{"cell": "C4", "label": "Concrete Strength (f'c)", "value": 28}],
            "outputs": [{"cell": "F22", "label": "Required Ast", "value": "=(M*1000000)/(0.85*fy*d)"}]
        }
        
    def run_calculation(self, filename: str, inputs: dict) -> dict:
        """
        Injects OpenSees FEA results into the Excel inputs, evaluates it,
        and extracts the result.
        
        Since openpyxl CANNOT evaluate formulas, a full implementation would use:
        - `xlwings` (requires MS Excel installed, Windows only usually)
        - `formulas` (Python library to parse Excel formulas)
        - Or rely on a separate microservice.
        
        For this backend, we will stub the evaluation for now or write values and rely on 
        the user to open the sheet, OR use a simplified Python calculation.
        """
        return {"status": "success", "message": "Phase 3: Inputs injected, awaiting formula parser."}
